import xlrd
import openpyxl
import urllib.request
import urllib.error
import urllib
import re
import random
import time

path1 = './Resources/Part1_Decrease_TranslateByTranslateGoogleAPI.xlsx'
workbook = xlrd.open_workbook(path1)    # 打开文件目标xlsx
Data_sheet = workbook.sheets()[0]   # 打开第一个表格：Sheet1-decrease
cols = Data_sheet.col_values(7)     # 打开表格第七列
del cols[0]  # 删除第七列第一行：译文
del cols[0]  # 删除第七列第二行：生化名称

path2 = 'Output.xlsx'
sheet_name = 'Output'
workbook = openpyxl.Workbook()
OutputSheet = workbook.active
OutputSheet.title = sheet_name

for i in range(19, 30):

    print(cols[i])
    keyword = urllib.parse.quote(cols[i])   # 字符转化
    # print(keyword)      # Debug Info
    url = 'http://www.molbase.cn/new/product/?keyword=' + str(keyword)
    print(url)      # Debug Info 信息获取资源定位URL
    # cnt = random.randint(0,10)
    headers = ("User-Agent", "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)")
    opener = urllib.request.build_opener()
    opener.addheaders = [headers]
    data = opener.open(url).read()
    data = str(data)

    req = r'<a target="_blank" href="(.*?)" class="detial-btn cut-2">'
    res = re.compile(req).findall(data)
    if not res:
        print('None')
        res.insert(0, 'None')

    surl = "http:" + res[0]

    OutputSheet.cell(row=i + 1, column=1, value=str(cols[i]))
    OutputSheet.cell(row=i + 1, column=2, value=str(surl))
    OutputSheet.cell(row=i + 1, column=3, value='1')
    print(str(cols[i]), str(surl))
    workbook.save(path2)
    time.sleep(random.randint(3, 30))
